"""Helper to read filings."""

from __future__ import annotations

from functools import cached_property
import json
import logging
import os
from pathlib import Path
import time

from arelle import PluginManager
from arelle.ModelDtsObject import ModelRelationship
from arelle.ModelRelationshipSet import ModelRelationshipSet
from arelle.ModelValue import QName
from arelle.ModelXbrl import ModelXbrl
from arelle.XbrlConst import parentChild, summationItem
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from pyesef.parse_xbrl_file.load_statement_definition import (
    StatementName,
    UpdateStatementDefinitionJson,
)
from pyesef.utils.data_management import asdict_with_properties

from ..const import PATH_PROJECT_ROOT
from ..error import PyEsefError
from .common import Controller, EsefData, clean_linkrole, load_model_xbrl
from .extract_definitions_to_csv import extract_definitions_to_csv
from .hierarchy import Hierarchy
from .read_facts import facts_to_data_list

FILE_ENDING_ZIP = ".zip"

PATH_FAILED = os.path.abspath(os.path.join(PATH_PROJECT_ROOT, "error"))
PATH_ARCHIVES = os.path.abspath(os.path.join(PATH_PROJECT_ROOT, "archives"))
PATH_PARSED = os.path.abspath(os.path.join(PATH_PROJECT_ROOT, "parsed"))
PATH_FAILED = os.path.abspath(os.path.join(PATH_PROJECT_ROOT, "error"))


def data_list_to_clean_df(data_list: list[EsefData]) -> pd.DataFrame:
    """Convert a list of filing data to a Pandas dataframe."""
    data_frame_from_data_class = pd.json_normalize(  # type: ignore[arg-type]
        asdict_with_properties(obj) for obj in data_list
    )

    if data_frame_from_data_class.empty:
        return pd.DataFrame()

    data_frame_from_data_class["period_end"] = pd.to_datetime(
        data_frame_from_data_class["period_end"]
    )

    # Make the value column an int
    data_frame_from_data_class["value"] = data_frame_from_data_class["value"].astype(
        int
    )

    # Drop beginning-of-year items
    data_frame_from_data_class = data_frame_from_data_class.query(
        "not (period_end.dt.month == 1 & period_end.dt.day == 1)"
    )

    # Drop items before 2020
    data_frame_from_data_class = data_frame_from_data_class.query(
        "period_end.dt.year > 2020"
    )

    # Drop zero values
    data_frame_from_data_class = data_frame_from_data_class.query("value != 0")

    df_before_duplicate_drop = data_frame_from_data_class.copy()

    # It's easier to look for duplicates when using ints than when using floats
    try:
        df_before_duplicate_drop.loc[:, "value_int"] = (
            df_before_duplicate_drop["value"] * 100
        ).astype(int)
    except OverflowError:
        df_before_duplicate_drop.loc[:, "value_int"] = (
            df_before_duplicate_drop["value"]
        ).astype(int)

    # Drop any duplicates
    df_before_duplicate_drop = df_before_duplicate_drop.drop_duplicates(
        subset=[
            "period_end",
            "lei",
            "wider_anchor_or_xml_name",
            "xml_name",
            "level_1",
            "value",
        ],
        ignore_index=True,
    )

    df_before_duplicate_drop = df_before_duplicate_drop.drop(columns=["value_int"])

    return df_before_duplicate_drop


def _extract_model_roles(
    model_xbrl: ModelXbrl,
) -> tuple[dict[str, str], Hierarchy]:
    """
    Extract a lookup table between XML item name and the item's role.

    This allows us to determine what financial statement an item belongs to, eg income
    statement, cash flow analysis or balance sheet.
    """
    to_model_to_linkrole_map: dict[str, str] = {}
    hierarchy_dict: dict[str, str] = {}

    rel_set: ModelRelationshipSet = model_xbrl.relationshipSet(summationItem)
    concepts_by_roles: dict[str, list[str]] = {}

    model_relationships: list[ModelRelationship] = rel_set.modelRelationships

    try:
        for rel in model_relationships:
            link = concepts_by_roles.get(rel.linkrole, [])

            from_clark_qname: QName = rel.fromModelObject.qname
            to_clark_qname: QName = rel.toModelObject.qname
            from_clark = from_clark_qname.clarkNotation
            to_clark = to_clark_qname.clarkNotation

            if to_clark_qname not in to_model_to_linkrole_map:
                to_model_to_linkrole_map[to_clark_qname.localName] = clean_linkrole(
                    rel.linkrole
                )

            if "summation-item" in rel.arcrole:
                # Create a lookup table of the parent item of each statement item, eg
                # DepreciationAndAmortisationExpense is part of OperatingExpense
                hierarchy_dict[to_clark_qname.localName] = from_clark_qname.localName

            if from_clark not in link and from_clark is not None:
                link.append(from_clark)

            if to_clark not in link and to_clark is not None:
                link.append(to_clark)

            if rel.linkrole not in concepts_by_roles:
                concepts_by_roles[rel.linkrole] = link

    except Exception as exc:
        raise PyEsefError("Unable to load model roles due to ", exc) from exc

    hierarchy_data = Hierarchy(hierarchy_dict=hierarchy_dict)

    return to_model_to_linkrole_map, hierarchy_data


class ReadFiling:
    """
    Read and save filings.

    The data will be stored in a Excel file.
    """

    TEMPLATE_OUTPUT_PATH_EXCEL = "output.xlsx"

    def __init__(
        self,
        filing_folder: str = PATH_ARCHIVES,
        should_move_parsed_file: bool = True,
    ) -> None:
        """Init class."""
        start_time = time.time()

        self.filing_folder = filing_folder
        self.file_to_parse_list: list[str] = []
        self.filing_list: list[EsefData] = []
        self.should_move_parsed_file = should_move_parsed_file
        self.output_df: pd.DataFrame
        self.definitions: pd.DataFrame = pd.DataFrame()

        self.cntlr = Controller()  # The Arelle controller

        # Add support for reading ESEF-files
        PluginManager.addPluginModule("validate/ESEF")

        self.find_files()
        self.parse_file_list()
        self.filings_to_clean_df()
        self.save_to_excel()

        # Close the controller
        self.cntlr.close()
        end_time = time.time()
        total_time = round(end_time - start_time, 0)
        self.cntlr.addToLog(
            f"Parsed {len(self.file_to_parse_list)} files in {total_time}s"
        )

    def find_files(self) -> None:
        """Loop through the archive folder and locate relevant files to parse."""
        for subdir, _, files in os.walk(PATH_ARCHIVES):
            for file in files:
                zip_file_path = subdir + os.sep + file

                if not zip_file_path.endswith(FILE_ENDING_ZIP):
                    continue

                self.file_to_parse_list.append(zip_file_path)

    @cached_property
    def model_role_map(self) -> dict[str, set[str]]:
        """Return a map of statement types and their xml items."""
        with open(
            UpdateStatementDefinitionJson.PATH_JSON_MAP_FILE, encoding="UTF-8"
        ) as json_file:
            return json.loads(json_file.read())

    def find_link_role(self, model_xbrl: ModelXbrl, name: str) -> str:
        """Find model link roles for cash flow."""
        base_taxonomy_clarks = set(self.model_role_map[name])

        max_score = 0
        filer_role = ""
        for role in model_xbrl.roleTypes.keys():
            role_pres_rels = model_xbrl.relationshipSet(parentChild, role)
            role_concept_clarks = {
                rel.toModelObject.qname.clarkNotation
                for rel in role_pres_rels.modelRelationships
            }
            for root in role_pres_rels.rootConcepts:
                role_concept_clarks.add(root.qname.clarkNotation)
            score = len(role_concept_clarks & base_taxonomy_clarks)
            if score > max_score:
                max_score = score
                filer_role = role

        clean_role = filer_role.split("/")[-1]

        return clean_role

    def parse_file_list(self) -> None:
        """PARSE FILE."""
        for idx, zip_file_path in enumerate(self.file_to_parse_list):
            try:
                # Load zip-file into a ModelXbrl instance
                model_xbrl = load_model_xbrl(
                    zip_file_path=zip_file_path,
                    cntlr=self.cntlr,
                )

                cash_flow_name = self.find_link_role(
                    model_xbrl=model_xbrl,
                    name=StatementName.CASH_FLOW.value,
                )
                income_statement_name = self.find_link_role(
                    model_xbrl=model_xbrl,
                    name=StatementName.INCOME_STATEMENT.value,
                )
                balance_sheet_name = self.find_link_role(
                    model_xbrl=model_xbrl,
                    name=StatementName.BALANCE_SHEET.value,
                )
                changes_equity_name = self.find_link_role(
                    model_xbrl=model_xbrl,
                    name=StatementName.CHANGES_EQUITY.value,
                )

                if self.definitions.empty and len(model_xbrl.facts):
                    self.definitions = extract_definitions_to_csv(
                        model_xbrl.facts[0].concept
                    )

                # Extract the model roles
                to_model_to_linkrole_map, _ = _extract_model_roles(
                    model_xbrl=model_xbrl,
                )

                fact_list = facts_to_data_list(
                    model_xbrl=model_xbrl,
                    to_model_to_linkrole_map=to_model_to_linkrole_map,
                    cash_flow_name=cash_flow_name,
                    income_statement_name=income_statement_name,
                    balance_sheet_name=balance_sheet_name,
                    changes_equity_name=changes_equity_name,
                )
                self.filing_list.extend(fact_list)
                model_xbrl.modelManager.cntlr.addToLog(
                    f"Finished working on: {idx}/{len(self.file_to_parse_list)}"
                )
                model_xbrl.close()

                # Move the filing folder to another location.
                # This helps us if the script stops due to memory
                # constraints.
                if self.should_move_parsed_file:
                    self.move_parsed_file(
                        zip_file_path=zip_file_path, target_path=PATH_PARSED
                    )
                    self.cntlr.addToLog("Moved files to parsed folder")

            except Exception as exc:
                if self.should_move_parsed_file:
                    self.move_parsed_file(
                        zip_file_path=zip_file_path, target_path=PATH_FAILED
                    )
                    self.cntlr.addToLog(
                        f"Moved file to error folder due to {exc}",
                        level=logging.WARNING,
                    )

    def filings_to_clean_df(self) -> None:
        """Return a clean df."""
        self.output_df = data_list_to_clean_df(self.filing_list)

    def save_to_excel(self) -> None:
        """Save file to Excel."""
        if self.output_df.empty:
            self.cntlr.addToLog(
                "Empty output dataframe. Output file not saved.", level=logging.WARNING
            )
            return

        with pd.ExcelWriter(
            self.TEMPLATE_OUTPUT_PATH_EXCEL,
            engine="openpyxl",
        ) as writer:
            self.output_df.to_excel(
                writer,
                index=False,
                sheet_name="Data",
                freeze_panes=(1, 0),
            )

            # Define a named style with the desired date format
            date_style = NamedStyle(name="date_style")
            date_style.number_format = "yyyy-mm-dd"

            # Get the workbook and sheet objects
            worksheet: Worksheet = writer.sheets["Data"]

            # Enable autofilter
            worksheet.auto_filter.ref = (
                "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
            )

            # Apply the date style to the entire column
            for cell in worksheet["A"]:
                cell.style = date_style

            # Define a named style for integer formatting
            int_style = NamedStyle(name="int_style")
            int_style.number_format = "0"

            # Apply the integer style to the value column (column 'B')
            for cell in worksheet["G"]:
                cell.style = int_style

            # Store the definitions in the Excel file
            if not self.definitions.empty:
                self.definitions.to_excel(
                    writer,
                    index=False,
                    sheet_name="Definitions",
                    freeze_panes=(1, 0),
                )

    @staticmethod
    def move_parsed_file(zip_file_path: str, target_path: str) -> None:
        """Move a file from the filings folder to the parsed folder."""
        Path(target_path).mkdir(parents=True, exist_ok=True)

        os.replace(
            zip_file_path,
            os.path.join(target_path, os.path.basename(zip_file_path)),
        )
